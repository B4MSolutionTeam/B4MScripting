

import re,os
from tkinter import Image
from xml.etree.ElementTree import PI
import win32com.client
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill
import pandas
#############       Create Data      #############
# outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
outlook = win32com.client.Dispatch("Outlook.Application")

outlook_output = outlook.GetNameSpace("MAPI")
inbox = outlook_output.Folders.Item(1)
a = inbox.Folders("TT_CLICK").Items
# Create WorkBook Excel
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.merge_cells('A1:D1')
worksheet['A1'] = "TT_CLICK Report"
worksheet['A1'].alignment = Alignment(horizontal='center')
worksheet['A1'].font = Font(size=20, bold=True)

cell_date_num = 2
cell_title = 3
for i in a:
    
    subject_string = i.Subject.split(" - ") 
    worksheet['A{}'.format(str(cell_date_num))] = "Date Report: "
    worksheet['B{}'.format(str(cell_date_num))] = str(subject_string[2])
    body_string = i.body.split("\n")
    df = pandas.read_excel('System List.xlsx', sheet_name=0)
    mylist = df["List of System"].tolist()
    for string_line in body_string:
        if re.search("[0-9]-", string_line):
            string_line_list = string_line.split("-")
            for system in mylist:
                if system in string_line_list[1].strip():
                    cell_date_num = cell_date_num + 1
                    worksheet["A{}".format(str(cell_date_num))] = str(string_line_list[1].strip())
                    cell_date_num = cell_date_num + 1
                    break
                
        if re.search("^[*]", string_line):
            string_line_list = string_line.split("\t")
            worksheet["A{}".format(str(cell_date_num))] = string_line_list[1].rpartition(":")[0]
            if "OK" in string_line_list[1].rpartition(":")[2].replace(" ",""):
                worksheet["B{}".format(str(cell_date_num))] = string_line_list[1].rpartition(":")[2]
                fill_cell = PatternFill(patternType='solid', fgColor='00ff00')
                worksheet["B{}".format(str(cell_date_num))].fill = fill_cell
            else:
                worksheet["B{}".format(str(cell_date_num))] = string_line_list[1].rpartition(":")[2]
                fill_cell = PatternFill(patternType='solid', fgColor='ff0000')
                worksheet["B{}".format(str(cell_date_num))].fill = fill_cell
            cell_date_num = cell_date_num + 1
            for idx, col in enumerate(worksheet.columns, 1):
                worksheet.column_dimensions[get_column_letter(idx)].auto_size = True
    cell_date_num = cell_date_num + 1
workbook.save("TT_CLICK_Report.xls")


############## Send Mail ##################
message = outlook.CreateItem(0)
message.To = "An.PhamNhat@vn.bosch.com"
message.Cc = "anh.nguyennam@vn.bosch.com"
message.Subject  = "Testing Send Email Outlook. TT Click Report"
message.Body = "This is testing email from Ta Thai Duy (CI/RDM11-VH) - TAD6HC. Please not response this email"
message.Attachments.Add(os.path.join(os.getcwd(),'TT_CLICK_Report.xls'))
#message.Display()
message.Send()
###################################################################

